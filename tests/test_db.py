from pathlib import Path

from openpyxl import load_workbook

from urwatcher.db import Database, resolve_sqlite_path
from urwatcher.models import DiffResult, Listing, Room


def test_resolve_sqlite_path_handles_relative(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    path = resolve_sqlite_path("sqlite:///./relative.db")
    assert path == tmp_path / "relative.db"


def test_database_initializes_schema(tmp_path):
    db_path = tmp_path / "urwatcher.db"
    db = Database(path=db_path)
    db.initialize()

    assert db_path.exists()
    assert db_path.stat().st_size > 0


def test_apply_changes_adds_and_removes_listings(tmp_path):
    db = Database(path=tmp_path / "urwatcher.db")
    db.initialize()

    listing = Listing(
        property_id="123",
        name="Sample",
        url="https://example.com/123.html",
        address="Sample Address",
        available_room_count=2,
    )
    diff_add = DiffResult(added=[listing], removed=[], unchanged=[])

    db.apply_listing_changes(executed_at="2025-01-01T00:00:00", diff=diff_add)

    records = db.fetch_listings(active_only=False)
    assert "123" in records
    assert records["123"].active is True
    assert records["123"].address == "Sample Address"
    assert records["123"].available_room_count == 2

    diff_remove = DiffResult(added=[], removed=[records["123"]], unchanged=[])
    db.apply_listing_changes(executed_at="2025-01-02T00:00:00", diff=diff_remove)

    updated = db.fetch_listings(active_only=False)["123"]
    assert updated.active is False


def test_apply_listing_changes_records_availability_change(tmp_path):
    db = Database(path=tmp_path / "availability.db")
    db.initialize()

    listing = Listing(
        property_id="123",
        name="Sample",
        url="https://example.com/123.html",
        address="Sample Address",
        available_room_count=1,
    )
    db.apply_listing_changes(
        executed_at="2025-01-01T00:00:00",
        diff=DiffResult(added=[listing], removed=[], unchanged=[]),
    )

    updated_listing = Listing(
        property_id="123",
        name="Sample",
        url="https://example.com/123.html",
        address="Sample Address",
        available_room_count=0,
    )
    db.apply_listing_changes(
        executed_at="2025-01-02T00:00:00",
        diff=DiffResult(added=[updated_listing], removed=[], unchanged=[]),
    )

    records = db.fetch_listings(active_only=False)
    assert records["123"].available_room_count == 0

    with db.connect() as conn:
        events = conn.execute(
            """
            SELECT event_type, details FROM listing_events
            WHERE property_id = ? ORDER BY id
            """,
            ("123",),
        ).fetchall()

    assert any(
        event_type == "availability_changed" and details == "1 -> 0"
        for event_type, details in events
    )


def test_unicode_listing_names_are_preserved(tmp_path):
    db = Database(path=tmp_path / "unicode.db")
    db.initialize()

    name = "千代田区"
    listing = Listing(
        property_id="jp-101",
        name=name,
        url="https://example.com/jp-101.html",
        address="千代田区千代田1-1",
        available_room_count=3,
    )
    diff = DiffResult(added=[listing], removed=[], unchanged=[])

    db.apply_listing_changes(executed_at="2025-02-01T00:00:00", diff=diff)

    stored = db.fetch_listings(active_only=True)["jp-101"]
    assert stored.name == name
    assert stored.address == "千代田区千代田1-1"


def test_apply_room_changes_adds_and_removes_rooms(tmp_path):
    db = Database(path=tmp_path / "rooms.db")
    db.initialize()

    # Seed listing for FK constraint.
    listing = Listing(
        property_id="P-1",
        name="Sample",
        url="https://example.com/property",
        address="1 Property Way",
        available_room_count=1,
    )
    db.apply_listing_changes(
        executed_at="2025-03-01T00:00:00",
        diff=DiffResult(added=[listing], removed=[], unchanged=[]),
    )

    room = Room(
        room_id="R-1",
        property_id="P-1",
        property_name="Sample",
        property_url="https://example.com/property",
        building_name="Building A",
        room_number="101",
        rent="60,000円",
        common_fee="3,000円",
        layout="2DK",
        floor_area="45㎡",
        floor="5階",
        room_url="https://example.com/property/rooms/101",
    )

    db.apply_room_changes(
        executed_at="2025-03-02T00:00:00",
        property_id="P-1",
        diff=DiffResult(added=[room], removed=[], unchanged=[]),
    )

    rooms = db.fetch_rooms(property_id="P-1", active_only=False)
    assert "R-1" in rooms
    assert rooms["R-1"].active is True

    db.apply_room_changes(
        executed_at="2025-03-03T00:00:00",
        property_id="P-1",
        diff=DiffResult(added=[], removed=[rooms["R-1"]], unchanged=[]),
    )

    updated = db.fetch_rooms(property_id="P-1", active_only=False)["R-1"]
    assert updated.active is False


def test_export_rooms_to_xlsx(tmp_path):
    db = Database(path=tmp_path / "export.db")
    db.initialize()

    listing = Listing(
        property_id="P-1",
        name="Sample",
        url="https://example.com/property",
        address="1 Property Way",
        available_room_count=1,
    )
    db.apply_listing_changes(
        executed_at="2025-04-01T00:00:00",
        diff=DiffResult(added=[listing], removed=[], unchanged=[]),
    )

    room = Room(
        room_id="R-1",
        property_id="P-1",
        property_name="Sample",
        property_url="https://example.com/property",
        building_name="Building A",
        room_number="101",
        rent="60,000円",
        common_fee="3,000円",
        layout="2DK",
        floor_area="45㎡",
        floor="5階",
        room_url="https://example.com/property/rooms/101",
    )
    db.apply_room_changes(
        executed_at="2025-04-01T00:05:00",
        property_id="P-1",
        diff=DiffResult(added=[room], removed=[], unchanged=[]),
    )

    export_path = tmp_path / "rooms.xlsx"
    db.export_rooms_to_xlsx(export_path)

    assert export_path.exists()
    workbook = load_workbook(export_path)
    worksheet = workbook.active
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    assert headers[:3] == ["address", "property_id", "room_id"]
    data_row = [cell.value for cell in next(worksheet.iter_rows(min_row=2, max_row=2))]
    assert data_row[0] == "1 Property Way"
    assert data_row[1] == "P-1"
    assert data_row[2] == "R-1"
